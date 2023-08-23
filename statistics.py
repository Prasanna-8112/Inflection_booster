import openpyxl,xlrd
from openpyxl import Workbook
import pathlib

file = pathlib.Path("StatusofUsers.xlsx")
if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet["A1"] = "Name"
    sheet["B1"] = "Courses"
    sheet["C1"] =  "Test"
    file.save("StatusofUsers.xlsx")
def Insert(name):
    file =  openpyxl.load_workbook("StatusofUsers.xlsx")
    sheet = file.active
    sheet.cell(column = 1,row = sheet.max_row + 1,value = name)
    sheet.cell(column = 2,row = sheet.max_row,value = 0)
    sheet.cell(column = 3,row = sheet.max_row,value = 0)
    file.save("StatusofUsers.xlsx")

def CheckName(name):
    file =  openpyxl.load_workbook("StatusofUsers.xlsx")
    sheet = file.active
    nameslist = sheet["A"]
    for cell in nameslist:
        if f'{cell.value}' == name:
            return f'{cell}'[15]
    return False


def GetValues(row):
    file =  openpyxl.load_workbook("StatusofUsers.xlsx")
    sheet = file.active
    cell1 =  "B" + str(row)
    cell2 = "C" + str(row)
    return [sheet[cell1].value,sheet[cell2].value]

def UpdateCourseSheet(row,count):
    file = openpyxl.load_workbook("StatusofUsers.xlsx")
    sheet = file.active
    cell = "B" + str(row)
    sheet[cell].value = count
    file.save("StatusofUsers.xlsx")


def UpdateTestSheet(row,count):
    file = openpyxl.load_workbook("StatusofUsers.xlsx")
    sheet = file.active
    cell = "C" + str(row)
    sheet[cell].value = count
    file.save("StatusofUsers.xlsx")

    
